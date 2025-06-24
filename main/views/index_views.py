from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm
from django.contrib.admin.views.decorators import staff_member_required
import openpyxl
from datetime import datetime
from main.models import Index, IndexValue
from io import BytesIO
from django.utils.timezone import localtime
from django.utils.dateformat import format as date_format
import time
from django.db import transaction


@login_required
@user_passes_test(lambda u: u.is_superuser)
def import_excel_view(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            return HttpResponse("❌ Aucun fichier n'a été envoyé.")

        try:
            # ✅ Utiliser openpyxl au lieu de pandas
            workbook = openpyxl.load_workbook(BytesIO(excel_file.read()))

            # Vérifier si la feuille "1. BDD" existe
            if "1. BDD" not in workbook.sheetnames:
                return HttpResponse("❌ Feuille '1. BDD' non trouvée dans l'Excel.")

            worksheet = workbook["1. BDD"]

            # Lire la première ligne pour les en-têtes
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(cell.value)
                else:
                    headers.append("")

            print("🔍 Colonnes détectées :", headers)

            if "Designation" not in headers:
                return HttpResponse("❌ Colonne 'Designation' manquante dans l'Excel.")

            # Trouver les indices des colonnes importantes
            designation_col = headers.index("Designation")
            unit_col = headers.index("Unit") if "Unit" in headers else None
            category_col = headers.index("category") if "category" in headers else None

            # Identifier les colonnes de dates
            date_columns = []
            for i, header in enumerate(headers):
                if isinstance(header, datetime) or (isinstance(header, str) and len(header) > 8):
                    try:
                        if isinstance(header, str):
                            parsed_date = datetime.strptime(header, "%Y-%m-%d")
                        else:
                            parsed_date = header
                        date_columns.append((i, parsed_date))
                    except:
                        continue

            # 🚀 NOUVEAU : Traitement par chunks pour éviter les timeouts
            total_rows = worksheet.max_row - 1  # -1 car on ignore la ligne d'en-tête
            chunk_size = 10  # Traiter 10 lignes à la fois
            processed_rows = 0
            start_time = time.time()

            print(f"📊 Total à traiter: {total_rows} lignes, par chunks de {chunk_size}")

            # Traiter chaque chunk
            for chunk_start in range(2, worksheet.max_row + 1, chunk_size):
                chunk_end = min(chunk_start + chunk_size, worksheet.max_row + 1)

                print(f"🔄 Traitement lignes {chunk_start} à {chunk_end-1}")

                # Utiliser une transaction pour chaque chunk (plus efficace)
                with transaction.atomic():

                    for row_num in range(chunk_start, chunk_end):
                        row = list(worksheet[row_num])

                        # Lire le nom de l'index
                        if designation_col < len(row) and row[designation_col].value:
                            index_name = str(row[designation_col].value).strip()
                        else:
                            continue

                        if not index_name:
                            continue

                        # Lire les champs associés
                        unit = "€"
                        if unit_col is not None and unit_col < len(row) and row[unit_col].value:
                            unit = str(row[unit_col].value)

                        category = "Autre"
                        if category_col is not None and category_col < len(row) and row[category_col].value:
                            category = str(row[category_col].value)

                        # ✅ Créer ou mettre à jour l'index
                        index_obj, created = Index.objects.update_or_create(
                            name=index_name,
                            defaults={
                                "unit": unit,
                                "category": category
                            }
                        )

                        # Traiter les valeurs pour chaque date
                        for col_index, date_value in date_columns:
                            if col_index < len(row) and row[col_index].value is not None:
                                raw_value = row[col_index].value

                                if raw_value is None or str(raw_value).strip() in ["", "-"]:
                                    continue

                                try:
                                    float_value = float(raw_value)
                                except (ValueError, TypeError):
                                    continue

                                # Créer ou mettre à jour la valeur
                                IndexValue.objects.update_or_create(
                                    index=index_obj,
                                    date=date_value.date(),
                                    defaults={"value": float_value}
                                )

                        processed_rows += 1

                # 🕐 Petite pause entre les chunks pour ne pas surcharger
                elapsed_time = time.time() - start_time
                print(f"⏱️ Chunk terminé. {processed_rows}/{total_rows} lignes. Temps écoulé: {elapsed_time:.1f}s")

                # Si on approche des 30 secondes, on arrête et on retourne un état
                if elapsed_time > 25:
                    return HttpResponse(f"⚠️ Importation partielle: {processed_rows}/{total_rows} lignes traitées. Relancez l'import pour continuer.")

            elapsed_time = time.time() - start_time
            return HttpResponse(f"✅ Importation terminée avec succès! {processed_rows} lignes traitées en {elapsed_time:.1f}s.")

        except openpyxl.utils.exceptions.InvalidFileException:
            return HttpResponse("❌ Fichier Excel invalide ou corrompu.")
        except Exception as e:
            return HttpResponse(f"❌ Erreur lors du traitement du fichier : {str(e)}")

    return render(request, "admin/import_excel.html")


@login_required
def liste_index_view(request):
    indexes = Index.objects.all()
    enriched_indexes = []

    for index in indexes:
        values = IndexValue.objects.filter(index=index).order_by('-date')
        if values.exists():
            latest_value = values.first()
            recent_values = list(values.order_by('date'))[-30:]
            value_data = {
                'date': latest_value.date.strftime("%d/%m/%Y"),
                'value': latest_value.value,
                'unit': index.unit,
                'dates': [date_format(v.date, "Y-m-d") for v in recent_values],
                'values': [v.value for v in recent_values],
            }
        else:
            value_data = {
                'date': "—",
                'value': "—",
                'unit': index.unit or "—",
                'history': [],
            }

        enriched_indexes.append({'index': index, 'latest': value_data})

    # 🔧 FIX de l'erreur 500: Vérifier si l'utilisateur est connecté
    context = {
        'indexes': enriched_indexes,
    }

    # Ajouter les favoris seulement si l'utilisateur est connecté et a un profil
    if request.user.is_authenticated:
        try:
            context['favorites'] = request.user.userprofile.favorite_indexes.all()
        except AttributeError:
            context['favorites'] = []
    else:
        context['favorites'] = []

    return render(request, 'liste_index.html', context)