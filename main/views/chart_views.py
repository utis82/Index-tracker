from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse 
from main.models import Index, IndexValue, Product, Part
from main.decorators import subscription_required
from django.core.serializers.json import DjangoJSONEncoder
from statistics import mean
from datetime import datetime, date as datetime_date
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required
@subscription_required
def index_viewer(request, index_id=None):
    """
    Vue principale pour l'analyse d'index avec comparaison de deux périodes
    MAINTENANT avec support des produits et pièces
    """
    user_profile = request.user.userprofile
    favorites = user_profile.favorite_indexes.all()

    # Récupération des index sélectionnés (maximum 3)
    selected_index_ids = request.GET.getlist("index_id")
    if not selected_index_ids and index_id:
        selected_index_ids = [str(index_id)]

    # Récupération des produits et pièces sélectionnés
    selected_product_ids = request.GET.getlist("product_id")
    selected_part_ids = request.GET.getlist("part_id")

    # Limiter le total à 3 éléments maximum
    total_selected = len(selected_index_ids) + len(selected_product_ids) + len(selected_part_ids)
    if total_selected > 3:
        # Tronquer pour garder seulement 3 éléments
        remaining = 3
        if remaining > 0 and selected_index_ids:
            keep_indexes = min(remaining, len(selected_index_ids))
            selected_index_ids = selected_index_ids[:keep_indexes]
            remaining -= keep_indexes
        if remaining > 0 and selected_product_ids:
            keep_products = min(remaining, len(selected_product_ids))
            selected_product_ids = selected_product_ids[:keep_products]
            remaining -= keep_products
        if remaining > 0 and selected_part_ids:
            keep_parts = min(remaining, len(selected_part_ids))
            selected_part_ids = selected_part_ids[:keep_parts]

    # Récupération des objets sélectionnés
    selected_indexes = favorites.filter(id__in=selected_index_ids)
    selected_products = Product.objects.filter(id__in=selected_product_ids, user=request.user)
    selected_parts = Part.objects.filter(id__in=selected_part_ids, product__user=request.user)

    # Récupération des paramètres de dates
    start_a = request.GET.get("start_a")
    end_a = request.GET.get("end_a")
    start_b = request.GET.get("start_b")
    end_b = request.GET.get("end_b")

    # Préparation des données pour le graphique
    series_data = []
    period_stats = {}

    # Récupérer les données d'index pour les calculs
    try:
        from main.utils import get_user_index_data
        index_data = get_user_index_data(request.user)
    except:
        index_data = {}

    # === AJOUT : Gestion des INDEX ===
    if selected_indexes.exists():
        for idx in selected_indexes:
            values = IndexValue.objects.filter(index=idx).order_by("date")
            if values.exists():
                dates = [v.date.strftime("%Y-%m-%d") for v in values]
                vals = [float(v.value) for v in values]
                series_data.append({
                    "name": f"IDX {idx.name}",
                    "id": f"index_{idx.id}",
                    "dates": dates,
                    "values": vals,
                    "unit": idx.unit or "Index",
                })

    # === NOUVEAU : Gestion des PRODUITS ===
    if selected_products.exists():
        for product in selected_products:
            data_points = []
            parts = product.parts.all()

            # Vérifier qu'il y a des tranches
            all_slices = []
            for part in parts:
                all_slices.extend(part.slices.all())

            if all_slices:
                # Récupérer tous les index utilisés dans ce produit
                index_ids = [
                    sl.index_id
                    for part in parts
                    for sl in part.slices.all()
                    if sl.component_type == "indexed" and sl.index_id
                ]

                # Vérifier s'il y a des tranches fixes
                has_fixed_slices = any(
                    sl.component_type == "fixed" 
                    for part in parts 
                    for sl in part.slices.all()
                )

                if index_ids:  # Il y a des tranches indexées
                    all_dates = sorted(set().union(*(index_data.get(i, {}).keys() for i in index_ids)))
                    product_reference_date = product.reference_date

                    for date in all_dates:
                        if date >= product_reference_date:
                            total_price = 0
                            # Calculer le prix de chaque pièce
                            for part in parts:
                                part_current_price = calculate_part_price_at_date(part, date, index_data)
                                total_price += part_current_price

                            data_points.append({
                                "date": date.strftime("%Y-%m-%d"),
                                "value": round(total_price, 2)
                            })
                elif has_fixed_slices:  # Seulement des tranches fixes
                    # Ligne horizontale
                    total_price = sum(calculate_part_price_at_date(part, product.reference_date, index_data) for part in parts)

                    # Points mensuels de la référence à aujourd'hui
                    ref_date = product.reference_date
                    today = datetime_date.today()
                    current_date = datetime_date(ref_date.year, ref_date.month, 1)

                    while current_date <= today:
                        data_points.append({
                            "date": current_date.strftime("%Y-%m-%d"),
                            "value": round(total_price, 2)
                        })

                        # Mois suivant
                        if current_date.month == 12:
                            current_date = datetime_date(current_date.year + 1, 1, 1)
                        else:
                            current_date = datetime_date(current_date.year, current_date.month + 1, 1)

                if data_points:
                    dates = [point["date"] for point in data_points]
                    values = [point["value"] for point in data_points]
                    series_data.append({
                        "name": f"PRD {product.name}",
                        "id": f"product_{product.id}",
                        "dates": dates,
                        "values": values,
                        "unit": "€",
                    })

    # === NOUVEAU : Gestion des PIÈCES ===
    if selected_parts.exists():
        for part in selected_parts:
            part_data_points = []

            # Vérifier qu'il y a des tranches
            if part.slices.exists():
                # Récupérer tous les index utilisés dans cette pièce
                part_index_ids = [
                    sl.index_id
                    for sl in part.slices.all()
                    if sl.component_type == "indexed" and sl.index_id
                ]

                # Vérifier s'il y a des tranches fixes
                has_fixed_slices = any(
                    sl.component_type == "fixed" 
                    for sl in part.slices.all()
                )

                if part_index_ids:  # La pièce a des tranches indexées
                    all_dates = sorted(set().union(*(index_data.get(i, {}).keys() for i in part_index_ids)))
                    part_reference_date = part.reference_date

                    for date in all_dates:
                        if date >= part_reference_date:
                            part_current_price = calculate_part_price_at_date(part, date, index_data)
                            part_data_points.append({
                                "date": date.strftime("%Y-%m-%d"),
                                "value": round(part_current_price, 2)
                            })
                elif has_fixed_slices:  # Seulement des tranches fixes
                    part_current_price = calculate_part_price_at_date(part, part.reference_date, index_data)

                    # Points mensuels de la référence à aujourd'hui
                    ref_date = part.reference_date
                    today = datetime_date.today()
                    current_date = datetime_date(ref_date.year, ref_date.month, 1)

                    while current_date <= today:
                        part_data_points.append({
                            "date": current_date.strftime("%Y-%m-%d"),
                            "value": round(part_current_price, 2)
                        })

                        # Mois suivant
                        if current_date.month == 12:
                            current_date = datetime_date(current_date.year + 1, 1, 1)
                        else:
                            current_date = datetime_date(current_date.year, current_date.month + 1, 1)

                if part_data_points:
                    dates = [point["date"] for point in part_data_points]
                    values = [point["value"] for point in part_data_points]
                    series_data.append({
                        "name": f"PRT {part.name}",
                        "id": f"part_{part.id}",
                        "dates": dates,
                        "values": values,
                        "unit": "€",
                    })

    # Calcul des statistiques si toutes les dates sont fournies
    if all([start_a, end_a, start_b, end_b]):
        try:
            date_a_start = datetime.strptime(start_a, "%Y-%m-%d").date()
            date_a_end = datetime.strptime(end_a, "%Y-%m-%d").date()
            date_b_start = datetime.strptime(start_b, "%Y-%m-%d").date()
            date_b_end = datetime.strptime(end_b, "%Y-%m-%d").date()

            # Validation des dates
            if date_a_start >= date_a_end or date_b_start >= date_b_end:
                raise ValueError("Dates invalides")

            # Statistiques pour les INDEX
            for idx in selected_indexes:
                values = IndexValue.objects.filter(index=idx).order_by("date")
                values_a = [float(v.value) for v in values if date_a_start <= v.date <= date_a_end]
                values_b = [float(v.value) for v in values if date_b_start <= v.date <= date_b_end]

                if values_a and values_b:
                    mean_a = mean(values_a)
                    mean_b = mean(values_b)
                    variation = ((mean_b - mean_a) / mean_a) * 100 if mean_a != 0 else 0

                    period_stats[f"IDX {idx.name}"] = {
                        "min_a": min(values_a),
                        "max_a": max(values_a),
                        "mean_a": round(mean_a, 2),
                        "min_b": min(values_b),
                        "max_b": max(values_b),
                        "mean_b": round(mean_b, 2),
                        "delta": round(mean_b - mean_a, 2),
                        "unit": idx.unit or "",
                        "variation": round(variation, 2),
                        "count_a": len(values_a),
                        "count_b": len(values_b)
                    }

            # Statistiques pour les PRODUITS
            for product in selected_products:
                product_values_a = []
                product_values_b = []

                # Calculer les prix du produit pour chaque jour des périodes
                for serie in series_data:
                    if serie["id"] == f"product_{product.id}":
                        dates = [datetime.strptime(d, "%Y-%m-%d").date() for d in serie["dates"]]
                        values = serie["values"]

                        for i, date in enumerate(dates):
                            if date_a_start <= date <= date_a_end:
                                product_values_a.append(values[i])
                            if date_b_start <= date <= date_b_end:
                                product_values_b.append(values[i])
                        break

                if product_values_a and product_values_b:
                    mean_a = mean(product_values_a)
                    mean_b = mean(product_values_b)
                    variation = ((mean_b - mean_a) / mean_a) * 100 if mean_a != 0 else 0

                    period_stats[f"PRD {product.name}"] = {
                        "min_a": min(product_values_a),
                        "max_a": max(product_values_a),
                        "mean_a": round(mean_a, 2),
                        "min_b": min(product_values_b),
                        "max_b": max(product_values_b),
                        "mean_b": round(mean_b, 2),
                        "delta": round(mean_b - mean_a, 2),
                        "unit": "€",
                        "variation": round(variation, 2),
                        "count_a": len(product_values_a),
                        "count_b": len(product_values_b)
                    }

            # Statistiques pour les PIÈCES
            for part in selected_parts:
                part_values_a = []
                part_values_b = []

                # Calculer les prix de la pièce pour chaque jour des périodes
                for serie in series_data:
                    if serie["id"] == f"part_{part.id}":
                        dates = [datetime.strptime(d, "%Y-%m-%d").date() for d in serie["dates"]]
                        values = serie["values"]

                        for i, date in enumerate(dates):
                            if date_a_start <= date <= date_a_end:
                                part_values_a.append(values[i])
                            if date_b_start <= date <= date_b_end:
                                part_values_b.append(values[i])
                        break

                if part_values_a and part_values_b:
                    mean_a = mean(part_values_a)
                    mean_b = mean(part_values_b)
                    variation = ((mean_b - mean_a) / mean_a) * 100 if mean_a != 0 else 0

                    period_stats[f"PRT {part.name}"] = {
                        "min_a": min(part_values_a),
                        "max_a": max(part_values_a),
                        "mean_a": round(mean_a, 2),
                        "min_b": min(part_values_b),
                        "max_b": max(part_values_b),
                        "mean_b": round(mean_b, 2),
                        "delta": round(mean_b - mean_a, 2),
                        "unit": "€",
                        "variation": round(variation, 2),
                        "count_a": len(part_values_a),
                        "count_b": len(part_values_b)
                    }

        except (ValueError, TypeError) as e:
            print(f"❌ Erreur de calcul des périodes : {e}")
            period_stats = {}

    context = {
        "user_favorites": favorites,
        "selected_indexes": selected_indexes,
        "selected_products": selected_products,  # NOUVEAU
        "selected_parts": selected_parts,        # NOUVEAU
        "selected_product_ids": selected_product_ids,  # NOUVEAU
        "selected_part_ids": selected_part_ids,        # NOUVEAU
        "series_data": json.dumps(series_data, cls=DjangoJSONEncoder),
        "period_stats": period_stats,
        "start_a": start_a,
        "end_a": end_a,
        "start_b": start_b,
        "end_b": end_b,
        "max_indexes": 3,
        "user_plan": user_profile.subscription_plan,
        "index_limit": user_profile.index_limit(),
        "user": request.user,  # Pour accéder aux produits/pièces dans le template
    }

    return render(request, "index_viewer.html", context)


def calculate_part_price_at_date(part, target_date, index_data):
    """Calcule le prix d'une pièce à une date donnée"""
    total_price = 0

    for slice_obj in part.slices.all():
        slice_reference_value = part.reference_price * (slice_obj.percentage / 100)

        if slice_obj.component_type == 'indexed' and slice_obj.index_id and slice_obj.percentage:
            # Calcul pour tranches indexées
            series = index_data.get(slice_obj.index_id, {})
            base_val = series.get(part.reference_date)
            current_val = series.get(target_date)

            if base_val and current_val and base_val != 0:
                evolution_ratio = current_val / base_val
                slice_current_value = slice_reference_value * evolution_ratio
                total_price += slice_current_value
            else:
                # Pas de données d'index, utiliser la valeur de référence
                total_price += slice_reference_value

        elif slice_obj.component_type == 'fixed' and slice_obj.percentage:
            # Calcul pour tranches fixes - reste constant
            total_price += slice_reference_value

    return total_price


@login_required
def toggle_favorite_ajax(request, index_id):
    """Vue AJAX pour basculer le statut favori d'un index"""
    if request.method != 'POST':
        return JsonResponse({"error": "Méthode non autorisée"}, status=405)

    try:
        index = get_object_or_404(Index, id=index_id)
        user_profile = request.user.userprofile

        if index in user_profile.favorite_indexes.all():
            user_profile.favorite_indexes.remove(index)
            is_favorite = False
        else:
            if not user_profile.can_add_favorite():
                return JsonResponse({
                    "error": f"Limite d'index atteinte ({user_profile.index_limit()})",
                    "is_favorite": False
                }, status=400)

            user_profile.favorite_indexes.add(index)
            is_favorite = True

        return JsonResponse({
            "is_favorite": is_favorite,
            "favorites_count": user_profile.favorite_indexes.count(),
            "index_limit": user_profile.index_limit()
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def get_index_data_ajax(request, index_id):
    """Vue AJAX pour récupérer les données d'un index spécifique"""
    try:
        index = get_object_or_404(Index, id=index_id)
        user_profile = request.user.userprofile

        if index not in user_profile.favorite_indexes.all():
            return JsonResponse({"error": "Index non autorisé"}, status=403)

        values = IndexValue.objects.filter(index=index).order_by("date")

        data = {
            "id": index.id,
            "name": index.name,
            "unit": index.unit,
            "category": index.category,
            "dates": [v.date.strftime("%Y-%m-%d") for v in values],
            "values": [float(v.value) for v in values],
            "count": values.count(),
            "date_range": {
                "start": values.first().date.strftime("%Y-%m-%d") if values.exists() else None,
                "end": values.last().date.strftime("%Y-%m-%d") if values.exists() else None
            }
        }

        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def export_analysis_data(request):
    """Vue pour exporter les données d'analyse en JSON ou CSV"""
    if request.method != 'POST':
        return JsonResponse({"error": "Méthode non autorisée"}, status=405)

    try:
        selected_ids = request.POST.getlist("index_id")
        start_a = request.POST.get("start_a")
        end_a = request.POST.get("end_a")
        start_b = request.POST.get("start_b")
        end_b = request.POST.get("end_b")
        export_format = request.POST.get("format", "json")

        user_profile = request.user.userprofile
        selected_indexes = user_profile.favorite_indexes.filter(id__in=selected_ids[:3])

        export_data = {
            "analysis_date": datetime.now().isoformat(),
            "user": request.user.username,
            "periods": {
                "period_1": {"start": start_a, "end": end_a},
                "period_2": {"start": start_b, "end": end_b}
            },
            "indexes": []
        }

        if all([start_a, end_a, start_b, end_b]):
            date_a_start = datetime.strptime(start_a, "%Y-%m-%d").date()
            date_a_end = datetime.strptime(end_a, "%Y-%m-%d").date()
            date_b_start = datetime.strptime(start_b, "%Y-%m-%d").date()
            date_b_end = datetime.strptime(end_b, "%Y-%m-%d").date()

            for idx in selected_indexes:
                values = IndexValue.objects.filter(index=idx).order_by("date")

                values_a = [v for v in values if date_a_start <= v.date <= date_a_end]
                values_b = [v for v in values if date_b_start <= v.date <= date_b_end]

                if values_a and values_b:
                    vals_a = [float(v.value) for v in values_a]
                    vals_b = [float(v.value) for v in values_b]

                    mean_a = mean(vals_a)
                    mean_b = mean(vals_b)
                    variation = ((mean_b - mean_a) / mean_a) * 100 if mean_a != 0 else 0

                    index_data = {
                        "name": idx.name,
                        "unit": idx.unit,
                        "category": idx.category,
                        "period_1": {
                            "min": min(vals_a),
                            "max": max(vals_a),
                            "mean": round(mean_a, 2),
                            "count": len(vals_a)
                        },
                        "period_2": {
                            "min": min(vals_b),
                            "max": max(vals_b),
                            "mean": round(mean_b, 2),
                            "count": len(vals_b)
                        },
                        "variation_percent": round(variation, 2)
                    }

                    export_data["indexes"].append(index_data)

        return JsonResponse(export_data)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def export_analysis_excel(request):
    """
    Vue pour exporter l'analyse complète vers Excel
    Utilise uniquement openpyxl (sans pandas)
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Méthode POST requise"}, status=405)

    try:
        # Récupération des paramètres
        selected_index_ids = request.POST.getlist("index_id")
        selected_product_ids = request.POST.getlist("product_id") 
        selected_part_ids = request.POST.getlist("part_id")

        start_a = request.POST.get("start_a")
        end_a = request.POST.get("end_a")
        start_b = request.POST.get("start_b")
        end_b = request.POST.get("end_b")

        # Validation des paramètres
        if not any([selected_index_ids, selected_product_ids, selected_part_ids]):
            return JsonResponse({"error": "Aucun élément sélectionné"}, status=400)

        # Récupération des objets
        user_profile = request.user.userprofile
        selected_indexes = user_profile.favorite_indexes.filter(id__in=selected_index_ids)
        selected_products = Product.objects.filter(id__in=selected_product_ids, user=request.user)
        selected_parts = Part.objects.filter(id__in=selected_part_ids, product__user=request.user)

        # Récupération des données d'index
        try:
            from main.utils import get_user_index_data
            index_data = get_user_index_data(request.user)
        except:
            index_data = {}

        # Générer les données pour l'export
        export_data = generate_export_data_excel(
            selected_indexes, selected_products, selected_parts,
            start_a, end_a, start_b, end_b, index_data
        )

        # Créer le workbook Excel avec openpyxl
        workbook = openpyxl.Workbook()

        # Supprimer la feuille par défaut
        workbook.remove(workbook.active)

        # Créer les feuilles
        create_summary_sheet_openpyxl(workbook, export_data, request.user)
        create_raw_data_sheet_openpyxl(workbook, export_data)

        if export_data['period_stats']:
            create_statistics_sheet_openpyxl(workbook, export_data)

        create_metadata_sheet_openpyxl(workbook, export_data)

        # Sauvegarder dans un BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Génération du nom de fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"analyse_indexes_{timestamp}.xlsx"

        # Réponse HTTP avec le fichier Excel
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        print(f"❌ Erreur export Excel: {e}")
        return JsonResponse({"error": f"Erreur lors de l'export: {str(e)}"}, status=500)


def generate_export_data_excel(selected_indexes, selected_products, selected_parts, 
                        start_a, end_a, start_b, end_b, index_data):
    """Génère toutes les données nécessaires pour l'export"""

    export_data = {
        'series_data': [],
        'period_stats': {},
        'metadata': {
            'export_date': datetime.now(),
            'periods': {
                'period_1': {'start': start_a, 'end': end_a},
                'period_2': {'start': start_b, 'end': end_b}
            },
            'elements': {
                'indexes_count': len(selected_indexes),
                'products_count': len(selected_products), 
                'parts_count': len(selected_parts)
            }
        }
    }

    # === DONNÉES DES INDEX ===
    for idx in selected_indexes:
        values = IndexValue.objects.filter(index=idx).order_by("date")
        if values.exists():
            dates = [v.date for v in values]
            vals = [float(v.value) for v in values]
            export_data['series_data'].append({
                "name": f"IDX {idx.name}",
                "type": "index",
                "id": f"index_{idx.id}",
                "dates": dates,
                "values": vals,
                "unit": idx.unit or "Index",
                "category": idx.category or "Non catégorisé",
                "metadata": {
                    "source": "Index",
                    "description": getattr(idx, 'description', ''),
                }
            })

    # === DONNÉES DES PRODUITS ===
    for product in selected_products:
        data_points = []
        parts = product.parts.all()

        all_slices = []
        for part in parts:
            all_slices.extend(part.slices.all())

        if all_slices:
            index_ids = [
                sl.index_id
                for part in parts
                for sl in part.slices.all()
                if sl.component_type == "indexed" and sl.index_id
            ]

            has_fixed_slices = any(
                sl.component_type == "fixed" 
                for part in parts 
                for sl in part.slices.all()
            )

            if index_ids:
                all_dates = sorted(set().union(*(index_data.get(i, {}).keys() for i in index_ids)))
                product_reference_date = product.reference_date

                for date in all_dates:
                    if date >= product_reference_date:
                        total_price = 0
                        for part in parts:
                            part_current_price = calculate_part_price_at_date(part, date, index_data)
                            total_price += part_current_price

                        data_points.append({
                            "date": date,
                            "value": round(total_price, 2)
                        })
            elif has_fixed_slices:
                total_price = sum(calculate_part_price_at_date(part, product.reference_date, index_data) for part in parts)

                ref_date = product.reference_date
                today = datetime_date.today()
                current_date = datetime_date(ref_date.year, ref_date.month, 1)

                while current_date <= today:
                    data_points.append({
                        "date": current_date,
                        "value": round(total_price, 2)
                    })

                    if current_date.month == 12:
                        current_date = datetime_date(current_date.year + 1, 1, 1)
                    else:
                        current_date = datetime_date(current_date.year, current_date.month + 1, 1)

            if data_points:
                dates = [point["date"] for point in data_points]
                values = [point["value"] for point in data_points]
                export_data['series_data'].append({
                    "name": f"PRD {product.name}",
                    "type": "product",
                    "id": f"product_{product.id}",
                    "dates": dates,
                    "values": values,
                    "unit": "€",
                    "category": "Produit",
                    "metadata": {
                        "source": "Produit calculé",
                        "reference_date": product.reference_date,
                        "parts_count": parts.count(),
                        "description": getattr(product, 'description', ''),
                    }
                })

    # === DONNÉES DES PIÈCES ===
    for part in selected_parts:
        part_data_points = []

        if part.slices.exists():
            part_index_ids = [
                sl.index_id
                for sl in part.slices.all()
                if sl.component_type == "indexed" and sl.index_id
            ]

            has_fixed_slices = any(
                sl.component_type == "fixed" 
                for sl in part.slices.all()
            )

            if part_index_ids:
                all_dates = sorted(set().union(*(index_data.get(i, {}).keys() for i in part_index_ids)))
                part_reference_date = part.reference_date

                for date in all_dates:
                    if date >= part_reference_date:
                        part_current_price = calculate_part_price_at_date(part, date, index_data)
                        part_data_points.append({
                            "date": date,
                            "value": round(part_current_price, 2)
                        })
            elif has_fixed_slices:
                part_current_price = calculate_part_price_at_date(part, part.reference_date, index_data)

                ref_date = part.reference_date
                today = datetime_date.today()
                current_date = datetime_date(ref_date.year, ref_date.month, 1)

                while current_date <= today:
                    part_data_points.append({
                        "date": current_date,
                        "value": round(part_current_price, 2)
                    })

                    if current_date.month == 12:
                        current_date = datetime_date(current_date.year + 1, 1, 1)
                    else:
                        current_date = datetime_date(current_date.year, current_date.month + 1, 1)

                if part_data_points:
                    dates = [point["date"] for point in part_data_points]
                    values = [point["value"] for point in part_data_points]
                    export_data['series_data'].append({
                        "name": f"PRT {part.name}",
                        "type": "part",
                        "id": f"part_{part.id}",
                        "dates": dates,
                        "values": values,
                        "unit": "€",
                        "category": "Pièce",
                        "metadata": {
                            "source": "Pièce calculée",
                            "reference_date": part.reference_date,
                            "reference_price": part.reference_price,
                            "product": part.product.name,
                            "description": getattr(part, 'description', ''),
                        }
                    })

    # === CALCUL DES STATISTIQUES DES PÉRIODES ===
    if all([start_a, end_a, start_b, end_b]):
        try:
            date_a_start = datetime.strptime(start_a, "%Y-%m-%d").date()
            date_a_end = datetime.strptime(end_a, "%Y-%m-%d").date()
            date_b_start = datetime.strptime(start_b, "%Y-%m-%d").date()
            date_b_end = datetime.strptime(end_b, "%Y-%m-%d").date()

            for serie in export_data['series_data']:
                values_a = []
                values_b = []

                for i, date in enumerate(serie['dates']):
                    if date_a_start <= date <= date_a_end:
                        values_a.append(serie['values'][i])
                    if date_b_start <= date <= date_b_end:
                        values_b.append(serie['values'][i])

                if values_a and values_b:
                    mean_a = mean(values_a)
                    mean_b = mean(values_b)
                    variation = ((mean_b - mean_a) / mean_a) * 100 if mean_a != 0 else 0

                    export_data['period_stats'][serie['name']] = {
                        "type": serie['type'],
                        "unit": serie['unit'],
                        "period_1": {
                            "min": round(min(values_a), 2),
                            "max": round(max(values_a), 2),
                            "mean": round(mean_a, 2),
                            "count": len(values_a)
                        },
                        "period_2": {
                            "min": round(min(values_b), 2),
                            "max": round(max(values_b), 2),
                            "mean": round(mean_b, 2),
                            "count": len(values_b)
                        },
                        "variation_percent": round(variation, 2),
                        "delta": round(mean_b - mean_a, 2)
                    }

        except Exception as e:
            print(f"❌ Erreur calcul statistiques: {e}")

    return export_data


def create_summary_sheet_openpyxl(workbook, export_data, user):
    """Crée la feuille de résumé avec openpyxl"""
    ws = workbook.create_sheet("Résumé")

    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    normal_font = Font(size=11)

    # En-têtes
    ws['A1'] = "Information"
    ws['B1'] = "Valeur"

    # Appliquer le style d'en-tête
    for cell in ['A1', 'B1']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = Alignment(horizontal="center")

    # Données
    data = [
        ['Date d\'export', export_data['metadata']['export_date'].strftime('%Y-%m-%d %H:%M:%S')],
        ['Utilisateur', user.username],
        ['Nombre d\'éléments analysés', len(export_data['series_data'])],
        ['Période 1', f"{export_data['metadata']['periods']['period_1']['start']} → {export_data['metadata']['periods']['period_1']['end']}"],
        ['Période 2', f"{export_data['metadata']['periods']['period_2']['start']} → {export_data['metadata']['periods']['period_2']['end']}"],
        ['', ''],
        ['Éléments inclus:', ''],
        ['- Index', export_data['metadata']['elements']['indexes_count']],
        ['- Produits', export_data['metadata']['elements']['products_count']],
        ['- Pièces', export_data['metadata']['elements']['parts_count']]
    ]

    for row_idx, (info, value) in enumerate(data, start=2):
        ws[f'A{row_idx}'] = info
        ws[f'B{row_idx}'] = value
        ws[f'A{row_idx}'].font = normal_font
        ws[f'B{row_idx}'].font = normal_font

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35


def create_raw_data_sheet_openpyxl(workbook, export_data):
    """Crée la feuille avec toutes les données brutes"""
    if not export_data['series_data']:
        return

    ws = workbook.create_sheet("Données brutes")

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    normal_font = Font(size=10)

    # Trouver toutes les dates uniques
    all_dates = set()
    for serie in export_data['series_data']:
        all_dates.update(serie['dates'])

    all_dates = sorted(all_dates)

    # En-têtes
    headers = ['Date']
    for serie in export_data['series_data']:
        headers.append(f"{serie['name']} ({serie['unit']})")

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Données
    for row_idx, date in enumerate(all_dates, start=2):
        # Date
        ws.cell(row=row_idx, column=1, value=date.strftime("%Y-%m-%d"))

        # Valeurs pour chaque série
        for col_idx, serie in enumerate(export_data['series_data'], start=2):
            # Créer un dictionnaire date->valeur pour cette série
            date_values = dict(zip(serie['dates'], serie['values']))
            value = date_values.get(date, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajuster la largeur des colonnes
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


def create_statistics_sheet_openpyxl(workbook, export_data):
    """Crée la feuille avec les statistiques des périodes"""
    ws = workbook.create_sheet("Statistiques périodes")

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    normal_font = Font(size=10)

    # En-têtes
    headers = [
        'Élément', 'Type', 'Unité',
        'P1 - Minimum', 'P1 - Maximum', 'P1 - Moyenne', 'P1 - Nb points',
        'P2 - Minimum', 'P2 - Maximum', 'P2 - Moyenne', 'P2 - Nb points',
        'Variation (%)', 'Delta absolu'
    ]

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Données
    row_idx = 2
    for element_name, stats in export_data['period_stats'].items():
        data_row = [
            element_name,
            stats['type'].title(),
            stats['unit'],
            stats['period_1']['min'],
            stats['period_1']['max'],
            stats['period_1']['mean'],
            stats['period_1']['count'],
            stats['period_2']['min'],
            stats['period_2']['max'],
            stats['period_2']['mean'],
            stats['period_2']['count'],
            stats['variation_percent'],
            stats['delta']
        ]

        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font

        row_idx += 1

    # Ajuster la largeur des colonnes
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


def create_metadata_sheet_openpyxl(workbook, export_data):
    """Crée la feuille avec les métadonnées des éléments"""
    ws = workbook.create_sheet("Métadonnées")

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    normal_font = Font(size=10)

    # En-têtes dynamiques selon les données disponibles
    base_headers = [
        'Nom', 'Type', 'Unité', 'Catégorie', 'Source',
        'Nombre de points', 'Date début', 'Date fin'
    ]

    # Ajouter des en-têtes spécifiques selon les types d'éléments
    additional_headers = []
    for serie in export_data['series_data']:
        if serie['type'] in ['product', 'part']:
            if 'Date de référence' not in additional_headers:
                additional_headers.append('Date de référence')
        if serie['type'] == 'product':
            if 'Nombre de pièces' not in additional_headers:
                additional_headers.append('Nombre de pièces')
        if serie['type'] == 'part':
            if 'Prix de référence' not in additional_headers:
                additional_headers.append('Prix de référence')
            if 'Produit parent' not in additional_headers:
                additional_headers.append('Produit parent')

    headers = base_headers + additional_headers + ['Description']

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Données
    for row_idx, serie in enumerate(export_data['series_data'], start=2):
        # Données de base
        data_row = [
            serie['name'],
            serie['type'].title(),
            serie['unit'],
            serie['category'],
            serie['metadata']['source'],
            len(serie['dates']),
            min(serie['dates']).strftime('%Y-%m-%d') if serie['dates'] else '',
            max(serie['dates']).strftime('%Y-%m-%d') if serie['dates'] else '',
        ]

        # Ajouter les données spécifiques
        for header in additional_headers:
            if header == 'Date de référence' and serie['type'] in ['product', 'part']:
                data_row.append(serie['metadata']['reference_date'].strftime('%Y-%m-%d'))
            elif header == 'Nombre de pièces' and serie['type'] == 'product':
                data_row.append(serie['metadata']['parts_count'])
            elif header == 'Prix de référence' and serie['type'] == 'part':
                data_row.append(serie['metadata']['reference_price'])
            elif header == 'Produit parent' and serie['type'] == 'part':
                data_row.append(serie['metadata']['product'])
            else:
                data_row.append('')

        # Description
        data_row.append(serie['metadata'].get('description', ''))

        # Écrire la ligne
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font

    # Ajuster la largeur des colonnes
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15